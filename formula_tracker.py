from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

def create_formula_based_tracker():
    """Create a formula-based Excel tracker with conditional formatting"""
    
    print("Creating formula-based Excel tracker...")
    
    # Start date: May 21, 2025
    start_date = datetime(2025, 5, 21)
    
    # Generate 3 months of dates (approximately 90 days)
    dates = []
    current = start_date
    for i in range(90):
        dates.append(current)
        current += timedelta(days=1)
    
    print(f"Generated {len(dates)} dates from {dates[0].strftime('%d-%m-%Y')} to {dates[-1].strftime('%d-%m-%Y')}")
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Task Tracker"
    
    # Headers
    ws['A1'] = "Date"
    ws['B1'] = "Task"
    ws['C1'] = "Status"
    
    # Instead of hardcoded dates, use Excel formulas
    # Start with the base date in A2, then use formulas for subsequent dates
    ws['A2'] = start_date.date()  # Base date: May 21, 2025
    
    # Use formulas for all subsequent dates
    for i in range(3, len(dates) + 2):
        ws[f'A{i}'] = f"=A{i-1}+1"  # Add 1 day to previous date
    
    print("Added formula-based dates to Excel")
    
    # Define styling
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    date_font = Font(bold=True, size=10)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    
    # Style headers
    for col in ['A1', 'B1', 'C1']:
        ws[col].font = header_font
        ws[col].fill = header_fill
        ws[col].border = thin_border
        ws[col].alignment = center_align
    
    # Style all data rows with borders and alignment
    for i in range(2, len(dates) + 2):
        # Style date cell (column A)
        ws[f'A{i}'].font = date_font
        ws[f'A{i}'].alignment = center_align
        ws[f'A{i}'].border = thin_border
        
        # Add borders to Task and Status columns
        ws[f'B{i}'].border = thin_border
        ws[f'C{i}'].border = thin_border
        ws[f'B{i}'].alignment = left_align
        ws[f'C{i}'].alignment = center_align
    
    print("Applied basic styling")
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    
    # Define conditional formatting colors
    weekend_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    today_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    
    # Create conditional formatting rules
    data_range = f"A2:C{len(dates) + 1}"
    
    # Rule 1: Highlight today's date (entire row)
    today_rule = FormulaRule(
        formula=[f'$A2=TODAY()'],
        fill=today_fill
    )
    
    # Rule 2: Highlight weekends (entire row)
    weekend_rule = FormulaRule(
        formula=[f'OR(WEEKDAY($A2,2)=6,WEEKDAY($A2,2)=7)'],
        fill=weekend_fill
    )
    
    # Apply conditional formatting rules
    # Note: Weekend rule is applied first, then today rule (so today overrides weekend if needed)
    ws.conditional_formatting.add(data_range, weekend_rule)
    ws.conditional_formatting.add(data_range, today_rule)
    
    print("Applied conditional formatting rules")
    
    # Add instructions below the data
    instruction_row = len(dates) + 4
    
    ws[f'A{instruction_row}'] = "INSTRUCTIONS:"
    ws[f'A{instruction_row}'].font = Font(bold=True, size=14, color="4472C4")
    
    instructions = [
        "• Enter your daily office tasks in the 'Task' column",
        "• Enter task status in the 'Status' column (e.g., Done, Pending, In Progress)",
        "• Gray rows automatically indicate weekends (formula-based)",
        "• Gold/yellow row automatically highlights TODAY's date (formula-based)",
        "• Dates are generated using Excel formulas starting from May 21, 2025",
        f"• This tracker covers {len(dates)} days from May 21, 2025",
        "• The highlighting updates automatically when you open the file on different dates"
    ]
    
    for idx, instruction in enumerate(instructions, start=1):
        ws[f'A{instruction_row + idx}'] = instruction
        ws[f'A{instruction_row + idx}'].font = Font(size=10)
    
    # Add legend
    legend_row = instruction_row + len(instructions) + 2
    
    ws[f'A{legend_row}'] = "LEGEND (Formula-Based):"
    ws[f'A{legend_row}'].font = Font(bold=True, size=12)
    
    ws[f'A{legend_row + 1}'] = "Weekend (Auto)"
    ws[f'B{legend_row + 1}'].fill = weekend_fill
    ws[f'B{legend_row + 1}'].border = thin_border
    
    ws[f'A{legend_row + 2}'] = "Today (Auto)"
    ws[f'B{legend_row + 2}'].fill = today_fill
    ws[f'B{legend_row + 2}'].border = thin_border
    
    # Add formula information
    formula_info_row = legend_row + 4
    ws[f'A{formula_info_row}'] = "FORMULA DETAILS:"
    ws[f'A{formula_info_row}'].font = Font(bold=True, size=12, color="4472C4")
    
    ws[f'A{formula_info_row + 1}'] = "• Dates: A2 contains May 21, 2025"
    ws[f'A{formula_info_row + 2}'] = "• A3 onwards: =A2+1 (adds 1 day to previous date)"
    ws[f'A{formula_info_row + 3}'] = "• Today highlighting: =A2=TODAY()"
    ws[f'A{formula_info_row + 4}'] = "• Weekend highlighting: =OR(WEEKDAY(A2,2)=6,WEEKDAY(A2,2)=7)"
    
    print("Added instructions and formula information")
    
    # Save the file
    filename = "Daily_Task_Tracker_Formula_Based_May2025.xlsx"
    wb.save(filename)
    
    print(f"✅ Formula-based Excel file created successfully!")
    print(f"📁 File name: {filename}")
    print(f"📅 Start date: May 21, 2025 (formula-based)")
    print(f"📝 Total rows: {len(dates)} dates")
    print(f"🎯 TODAY() function will automatically highlight current date")
    print(f"🔘 Weekends automatically detected using WEEKDAY() function")
    print(f"📊 All conditional formatting is formula-based and dynamic")
    
    return filename

if __name__ == "__main__":
    try:
        create_formula_based_tracker()
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
